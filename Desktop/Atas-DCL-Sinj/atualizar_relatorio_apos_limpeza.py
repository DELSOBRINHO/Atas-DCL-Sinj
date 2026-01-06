#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ATUALIZAR RELATÓRIO APÓS LIMPEZA DE DUPLICATAS
===============================================

Atualiza o relatório com os 346 DCLs únicos (após remoção de duplicatas)

Uso:
    python atualizar_relatorio_apos_limpeza.py

Autor: Sistema de Automação CLDF
Data: 2025-12-22
"""

import json
import re
from pathlib import Path

# ======================================================================
# CONFIGURAÇÕES
# ======================================================================

USUARIO = "omega"
DIR_DOWNLOADS = Path(f"C:/Users/{USUARIO}/Desktop/Atas-DCL-Sinj/downloads_2007")
DIR_LINKS = Path(f"C:/Users/{USUARIO}/Desktop/Atas-DCL-Sinj/links_2007")
ARQUIVO_LINKS = DIR_LINKS / "dcls_2007.json"
ARQUIVO_SAIDA_JSON = Path(f"C:/Users/{USUARIO}/Desktop/Atas-DCL-Sinj/relatorio_dcls_2007_2008_limpo.json")
ARQUIVO_SAIDA_XLSX = Path(f"C:/Users/{USUARIO}/Desktop/Atas-DCL-Sinj/relatorio_dcls_2007_2008_limpo.xlsx")

# ======================================================================
# FUNÇÕES PRINCIPAIS
# ======================================================================

def carregar_links():
    """Carrega links do arquivo JSON"""
    if not ARQUIVO_LINKS.exists():
        return {}
    
    try:
        with open(ARQUIVO_LINKS, 'r', encoding='utf-8') as f:
            links = json.load(f)
        
        links_dict = {}
        for link in links:
            try:
                partes = link['url'].split('DCL%20n%C2%BA%20')
                if len(partes) > 1:
                    numero = partes[1].split('%20')[0]
                    ano = link['ano']
                    mes = link['mes']
                    chave = f"DCL_{ano}-{mes:02d}-{numero:03d}"
                    links_dict[chave] = link['url']
            except:
                pass
        
        return links_dict
    
    except Exception as e:
        print(f"❌ Erro ao carregar links: {e}")
        return {}

def gerar_relatorio():
    """Gera relatório dos DCLs após limpeza"""
    print("\n" + "="*70)
    print("ATUALIZAR RELATÓRIO APÓS LIMPEZA DE DUPLICATAS")
    print("="*70)
    
    # Carregar links
    print("\n📂 CARREGANDO LINKS")
    links_dict = carregar_links()
    print(f"✅ {len(links_dict)} links carregados")
    
    # Listar todos os PDFs (após limpeza)
    print("\n📁 LISTANDO ARQUIVOS PDF (APÓS LIMPEZA)")
    todos_pdfs = sorted(DIR_DOWNLOADS.glob('*.pdf'))
    
    print(f"✅ {len(todos_pdfs)} PDFs encontrados")
    
    if not todos_pdfs:
        print("❌ Nenhum PDF encontrado")
        return
    
    # Gerar dados
    print("\n📊 GERANDO RELATÓRIO")
    print("="*70 + "\n")
    
    resultados = []
    
    for i, arquivo in enumerate(todos_pdfs, 1):
        nome = arquivo.stem
        partes = nome.split('-')
        
        ano = int(partes[1]) if len(partes) > 1 else 0
        mes = int(partes[2]) if len(partes) > 2 else 0
        tamanho_kb = arquivo.stat().st_size / 1024
        
        # Procurar URL
        chave = arquivo.name.replace('.pdf', '')
        url = links_dict.get(chave, '')
        
        resultado = {
            'numero': i,
            'arquivo': arquivo.name,
            'ano': ano,
            'mes': mes,
            'tamanho_kb': tamanho_kb,
            'status': 'processado',
            'url': url
        }
        
        resultados.append(resultado)
        
        if i % 50 == 0:
            print(f"[{i:3d}/{len(todos_pdfs)}] ✅ Processados")
    
    print(f"[{len(todos_pdfs):3d}/{len(todos_pdfs)}] ✅ Processados")
    
    # Salvar JSON
    print(f"\n💾 SALVANDO RESULTADOS")
    print("="*70)
    
    with open(ARQUIVO_SAIDA_JSON, 'w', encoding='utf-8') as f:
        json.dump(resultados, f, ensure_ascii=False, indent=2)
    
    print(f"✅ JSON salvo em: {ARQUIVO_SAIDA_JSON}")
    
    # Salvar Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "DCLs"
        
        # Cabeçalhos
        headers = ['Nº', 'Arquivo', 'Ano', 'Mês', 'Tamanho (KB)', 'Status', 'URL']
        ws.append(headers)
        
        # Estilo do cabeçalho
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Dados
        for resultado in resultados:
            ws.append([
                resultado.get('numero', ''),
                resultado.get('arquivo', ''),
                resultado.get('ano', ''),
                resultado.get('mes', ''),
                f"{resultado.get('tamanho_kb', 0):.0f}",
                resultado.get('status', ''),
                resultado.get('url', '')
            ])
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 50
        
        # Salvar
        wb.save(ARQUIVO_SAIDA_XLSX)
        
        print(f"✅ Excel salvo em: {ARQUIVO_SAIDA_XLSX}")
    
    except Exception as e:
        print(f"⚠️  Erro ao salvar Excel: {e}")
    
    # Estatísticas
    print(f"\n📊 ESTATÍSTICAS")
    print("="*70)
    
    total = len(resultados)
    tamanho_total_mb = sum(r.get('tamanho_kb', 0) for r in resultados) / 1024
    
    # Contar por mês
    por_mes = {}
    for r in resultados:
        chave = f"{r['ano']}-{r['mes']:02d}"
        por_mes[chave] = por_mes.get(chave, 0) + 1
    
    print(f"   Total:              {total}")
    print(f"   Tamanho Total:      {tamanho_total_mb:.1f} MB")
    print(f"\n   Distribuição por Período:")
    
    for chave in sorted(por_mes.keys()):
        print(f"      {chave}: {por_mes[chave]} DCLs")
    
    print(f"\n✅ RELATÓRIO ATUALIZADO COM SUCESSO!")

if __name__ == "__main__":
    gerar_relatorio()

