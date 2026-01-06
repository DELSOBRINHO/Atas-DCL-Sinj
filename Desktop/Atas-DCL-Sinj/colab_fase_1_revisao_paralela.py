"""
GOOGLE COLAB - FASE 1 REVISÃO E ATUALIZAÇÃO COM PROCESSAMENTO PARALELO
Processa DCLs em paralelo, revisa e atualiza o relatório Excel
"""

# ============================================================================
# INSTALAÇÃO DE DEPENDÊNCIAS (executar no Colab)
# ============================================================================

# !pip install -q pdfplumber openpyxl PyPDF2 python-magic-bin

# ============================================================================
# IMPORTS
# ============================================================================

import os
import json
import re
from pathlib import Path
from datetime import datetime
from concurrent.futures import ProcessPoolExecutor, as_completed
import pdfplumber
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd

# ============================================================================
# CONFIGURAÇÃO
# ============================================================================

# Para Colab: montar Google Drive
# from google.colab import drive
# drive.mount('/content/drive')
# DOWNLOADS_DIR = "/content/drive/My Drive/Atas-DCL-Sinj/downloads_2007"
# RELATORIO_ENTRADA = "/content/drive/My Drive/Atas-DCL-Sinj/relatorio_dcls_2007.xlsx"
# RELATORIO_SAIDA = "/content/drive/My Drive/Atas-DCL-Sinj/relatorio_dcls_2007_revisado.xlsx"

# Para execução local
DOWNLOADS_DIR = "downloads_2007"
RELATORIO_ENTRADA = "relatorio_dcls_2007.xlsx"
RELATORIO_SAIDA = "relatorio_dcls_2007_revisado.xlsx"
RELATORIO_JSON = "relatorio_dcls_2007_revisado.json"

# ============================================================================
# FUNÇÕES DE PROCESSAMENTO
# ============================================================================

def extrair_info_pdf(pdf_path):
    """Extrai informações de um PDF"""
    
    info = {
        "filename": os.path.basename(pdf_path),
        "tamanho_kb": os.path.getsize(pdf_path) / 1024,
        "paginas": 0,
        "tem_sumario": False,
        "tem_ata": False,
        "tipo_ata": None,
        "tipo_sessao": None,
        "numero_sessao": None,
        "data_sessao": None,
        "status": "Processado",
        "notas": []
    }
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            info["paginas"] = len(pdf.pages)
            
            # Processar primeiras 3 páginas
            for page_num in range(min(3, len(pdf.pages))):
                try:
                    page = pdf.pages[page_num]
                    text = page.extract_text()
                    
                    if not text:
                        continue
                    
                    text_upper = text.upper()
                    
                    # Detectar sumário
                    if "SUMÁRIO" in text_upper or "ÍNDICE" in text_upper:
                        info["tem_sumario"] = True
                    
                    # Detectar ata
                    if "ATA" in text_upper or "SESSÃO" in text_upper:
                        info["tem_ata"] = True
                    
                    # Detectar tipo de ata
                    if "CIRCUNSTANCIADA" in text_upper:
                        info["tipo_ata"] = "Circunstanciada"
                    elif "SUCINTA" in text_upper:
                        info["tipo_ata"] = "Sucinta"
                    
                    # Detectar tipo de sessão
                    if "EXTRAORDINÁRIA" in text_upper:
                        info["tipo_sessao"] = "Extraordinária"
                    elif "SOLENE" in text_upper:
                        info["tipo_sessao"] = "Solene"
                    elif "ORDINÁRIA" in text_upper:
                        info["tipo_sessao"] = "Ordinária"
                    
                    # Extrair número de sessão
                    pattern = r'(\d+)(?:ª|a)\s+(?:SESSÃO|SESSÃO)'
                    matches = re.findall(pattern, text_upper)
                    if matches and not info["numero_sessao"]:
                        info["numero_sessao"] = matches[0]
                    
                    # Extrair data
                    pattern_data = r'(\d{1,2})\s+de\s+(\w+)\s+de\s+(\d{4})'
                    matches_data = re.findall(pattern_data, text, re.IGNORECASE)
                    if matches_data and not info["data_sessao"]:
                        info["data_sessao"] = f"{matches_data[0][2]}-{matches_data[0][1]}-{matches_data[0][0]}"
                
                except Exception as e:
                    info["notas"].append(f"Erro página {page_num + 1}: {str(e)[:50]}")
        
        # Validações
        if not info["tem_ata"]:
            info["notas"].append("Nenhuma ata detectada")
        
        if not info["tipo_ata"]:
            info["notas"].append("Tipo de ata não identificado")
        
        if not info["tipo_sessao"]:
            info["notas"].append("Tipo de sessão não identificado")
        
        if not info["numero_sessao"]:
            info["notas"].append("Número de sessão não encontrado")
        
        if not info["data_sessao"]:
            info["notas"].append("Data de sessão não encontrada")
    
    except Exception as e:
        info["status"] = "Erro"
        info["notas"].append(f"Erro ao processar: {str(e)[:100]}")
    
    return info

def processar_dcls_paralelo(num_workers=4):
    """Processa DCLs em paralelo"""
    
    print(f"🔄 Processando {len(os.listdir(DOWNLOADS_DIR))} DCLs em paralelo...")
    print(f"   Workers: {num_workers}\n")
    
    dcl_files = sorted([f for f in os.listdir(DOWNLOADS_DIR) if f.endswith('.pdf')])
    resultados = []
    
    with ProcessPoolExecutor(max_workers=num_workers) as executor:
        futures = {
            executor.submit(extrair_info_pdf, os.path.join(DOWNLOADS_DIR, f)): f 
            for f in dcl_files
        }
        
        completed = 0
        for future in as_completed(futures):
            completed += 1
            resultado = future.result()
            resultados.append(resultado)
            
            # Mostrar progresso
            if completed % 10 == 0:
                print(f"   [{completed}/{len(dcl_files)}] Processados...")
    
    print(f"\n✅ {len(resultados)} DCLs processados com sucesso!\n")
    return resultados

def atualizar_relatorio_excel(resultados):
    """Atualiza relatório Excel com informações revisadas"""
    
    print("📊 Atualizando relatório Excel...\n")
    
    # Carregar relatório original
    wb = openpyxl.load_workbook(RELATORIO_ENTRADA)
    ws = wb.active
    
    # Adicionar colunas de revisão
    headers_novos = [
        "Páginas",
        "Tem Sumário",
        "Tem Ata",
        "Tipo Ata",
        "Tipo Sessão",
        "Nº Sessão",
        "Data Sessão",
        "Observações"
    ]
    
    # Encontrar coluna inicial para novos dados
    col_inicio = ws.max_column + 1
    
    # Adicionar cabeçalhos
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for idx, header in enumerate(headers_novos, col_inicio):
        cell = ws.cell(row=1, column=idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Preencher dados
    for row_idx, resultado in enumerate(resultados, 2):
        ws.cell(row=row_idx, column=col_inicio).value = resultado["paginas"]
        ws.cell(row=row_idx, column=col_inicio + 1).value = "Sim" if resultado["tem_sumario"] else "Não"
        ws.cell(row=row_idx, column=col_inicio + 2).value = "Sim" if resultado["tem_ata"] else "Não"
        ws.cell(row=row_idx, column=col_inicio + 3).value = resultado["tipo_ata"] or "N/A"
        ws.cell(row=row_idx, column=col_inicio + 4).value = resultado["tipo_sessao"] or "N/A"
        ws.cell(row=row_idx, column=col_inicio + 5).value = resultado["numero_sessao"] or "N/A"
        ws.cell(row=row_idx, column=col_inicio + 6).value = resultado["data_sessao"] or "N/A"
        ws.cell(row=row_idx, column=col_inicio + 7).value = "; ".join(resultado["notas"]) if resultado["notas"] else "OK"
    
    # Ajustar largura das colunas
    for idx in range(col_inicio, col_inicio + len(headers_novos)):
        ws.column_dimensions[get_column_letter(idx)].width = 18
    
    # Salvar
    wb.save(RELATORIO_SAIDA)
    print(f"✅ Relatório atualizado: {RELATORIO_SAIDA}\n")

def salvar_json(resultados):
    """Salva resultados em JSON"""
    
    print("💾 Salvando dados em JSON...\n")
    
    with open(RELATORIO_JSON, 'w', encoding='utf-8') as f:
        json.dump(resultados, f, indent=2, ensure_ascii=False)
    
    print(f"✅ Dados salvos: {RELATORIO_JSON}\n")

def gerar_estatisticas(resultados):
    """Gera estatísticas dos resultados"""
    
    print("📈 ESTATÍSTICAS DA REVISÃO\n")
    print("=" * 60)
    
    total = len(resultados)
    com_ata = sum(1 for r in resultados if r["tem_ata"])
    com_sumario = sum(1 for r in resultados if r["tem_sumario"])
    com_tipo_ata = sum(1 for r in resultados if r["tipo_ata"])
    com_tipo_sessao = sum(1 for r in resultados if r["tipo_sessao"])
    com_numero_sessao = sum(1 for r in resultados if r["numero_sessao"])
    com_data_sessao = sum(1 for r in resultados if r["data_sessao"])
    com_erro = sum(1 for r in resultados if r["status"] == "Erro")
    
    print(f"Total de DCLs processados:        {total}")
    print(f"DCLs com ata:                     {com_ata} ({com_ata/total*100:.1f}%)")
    print(f"DCLs com sumário:                 {com_sumario} ({com_sumario/total*100:.1f}%)")
    print(f"DCLs com tipo de ata:             {com_tipo_ata} ({com_tipo_ata/total*100:.1f}%)")
    print(f"DCLs com tipo de sessão:          {com_tipo_sessao} ({com_tipo_sessao/total*100:.1f}%)")
    print(f"DCLs com número de sessão:        {com_numero_sessao} ({com_numero_sessao/total*100:.1f}%)")
    print(f"DCLs com data de sessão:          {com_data_sessao} ({com_data_sessao/total*100:.1f}%)")
    print(f"DCLs com erro:                    {com_erro} ({com_erro/total*100:.1f}%)")
    print("=" * 60 + "\n")
    
    return {
        "total": total,
        "com_ata": com_ata,
        "com_sumario": com_sumario,
        "com_tipo_ata": com_tipo_ata,
        "com_tipo_sessao": com_tipo_sessao,
        "com_numero_sessao": com_numero_sessao,
        "com_data_sessao": com_data_sessao,
        "com_erro": com_erro
    }

# ============================================================================
# EXECUÇÃO PRINCIPAL
# ============================================================================

if __name__ == "__main__":
    print("\n" + "=" * 60)
    print("🚀 FASE 1 - REVISÃO E ATUALIZAÇÃO COM PROCESSAMENTO PARALELO")
    print("=" * 60 + "\n")
    
    # Processar DCLs em paralelo
    resultados = processar_dcls_paralelo(num_workers=4)
    
    # Atualizar relatório Excel
    atualizar_relatorio_excel(resultados)
    
    # Salvar JSON
    salvar_json(resultados)
    
    # Gerar estatísticas
    stats = gerar_estatisticas(resultados)
    
    print("✅ FASE 1 CONCLUÍDA COM SUCESSO!")
    print(f"\n📁 Arquivos gerados:")
    print(f"   - {RELATORIO_SAIDA}")
    print(f"   - {RELATORIO_JSON}")
    print(f"\n🎯 Próximo passo: Executar Fase 2 (Enriquecimento de Metadados)")

