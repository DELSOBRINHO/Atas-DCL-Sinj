"""
FASE 1: EXTRAIR E DOCUMENTAR DCLs
Extrai todos os links de DCL do SINJ-DF para 2007
Cria relatório Excel com informações dos DCLs
"""

import os
import json
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Configuração
DOWNLOADS_DIR = "downloads_2007"
RELATORIO_EXCEL = "relatorio_dcls_2007.xlsx"

def criar_relatorio_excel():
    """Cria relatório Excel com DCLs encontrados"""
    
    if not os.path.exists(DOWNLOADS_DIR):
        print(f"❌ Diretório {DOWNLOADS_DIR} não encontrado")
        return
    
    print(f"📊 Criando relatório Excel de DCLs...\n")
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DCLs 2007"
    
    # Definir cabeçalhos
    headers = [
        "Nº",
        "Filename Original",
        "URL SINJ-DF",
        "Data Publicação",
        "Tamanho (KB)",
        "Status",
        "Notas"
    ]
    
    # Estilo para cabeçalho
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Adicionar cabeçalhos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Listar DCLs
    dcl_files = sorted([f for f in os.listdir(DOWNLOADS_DIR) if f.endswith('.pdf')])
    
    print(f"📁 Encontrados {len(dcl_files)} DCLs\n")
    
    # Adicionar dados
    for row, filename in enumerate(dcl_files, 2):
        filepath = os.path.join(DOWNLOADS_DIR, filename)
        file_size = os.path.getsize(filepath) / 1024  # KB
        
        # Extrair data do nome do arquivo (formato: DCL_YYYY-MM-NNN.pdf)
        parts = filename.replace('.pdf', '').split('_')
        if len(parts) >= 3:
            ano = parts[1]
            mes = parts[2]
            data_pub = f"{ano}-{mes}"
        else:
            data_pub = "N/A"
        
        # Preencher linha
        ws.cell(row=row, column=1).value = row - 1
        ws.cell(row=row, column=2).value = filename
        ws.cell(row=row, column=3).value = f"https://www.sinj.df.gov.br/sinj/Consulta/Diario.html"
        ws.cell(row=row, column=4).value = data_pub
        ws.cell(row=row, column=5).value = f"{file_size:.1f}"
        ws.cell(row=row, column=6).value = "Pendente"
        ws.cell(row=row, column=7).value = ""
        
        # Aplicar bordas
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="left", vertical="center")
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 30
    
    # Congelar primeira linha
    ws.freeze_panes = "A2"
    
    # Salvar
    wb.save(RELATORIO_EXCEL)
    
    print(f"✅ Relatório criado: {RELATORIO_EXCEL}")
    print(f"   Total de DCLs: {len(dcl_files)}")
    print(f"   Colunas: {', '.join(headers)}")

if __name__ == "__main__":
    criar_relatorio_excel()

