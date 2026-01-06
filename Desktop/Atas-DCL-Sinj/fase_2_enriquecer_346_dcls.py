#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
FASE 2: ENRIQUECIMENTO DE METADADOS - 346 DCLs
===============================================

Extrai informações detalhadas de cada DCL:
- Data real da sessão
- Tipo de sessão (Ordinária, Extraordinária, etc.)
- Número da sessão
- Tipo de ata (Sucinta, Circunstanciada)
- Paginação

Entrada: relatorio_final_346_dcls.xlsx
Saída: relatorio_fase2_346_dcls_enriquecido.xlsx

Uso:
    python fase_2_enriquecer_346_dcls.py

Autor: Sistema de Automação CLDF
Data: 2025-12-22
"""

import json
import re
from pathlib import Path
from datetime import datetime

# ======================================================================
# CONFIGURAÇÕES
# ======================================================================

USUARIO = "omega"
DIR_DOWNLOADS = Path(f"C:/Users/{USUARIO}/Desktop/Atas-DCL-Sinj/downloads_2007")
ARQUIVO_ENTRADA = Path(f"C:/Users/{USUARIO}/Desktop/Atas-DCL-Sinj/relatorio_final_346_dcls.xlsx")
ARQUIVO_SAIDA_XLSX = Path(f"C:/Users/{USUARIO}/Desktop/Atas-DCL-Sinj/relatorio_fase2_346_dcls_enriquecido.xlsx")
ARQUIVO_SAIDA_JSON = Path(f"C:/Users/{USUARIO}/Desktop/Atas-DCL-Sinj/relatorio_fase2_346_dcls_enriquecido.json")

# ======================================================================
# FUNÇÕES PRINCIPAIS
# ======================================================================

def instalar_dependencias():
    """Instala dependências necessárias"""
    print("📦 Verificando dependências...")
    
    import subprocess
    
    pacotes = {
        "pdfplumber": "pdfplumber",
        "openpyxl": "openpyxl"
    }
    
    for modulo, pacote in pacotes.items():
        try:
            __import__(modulo)
            print(f"   ✅ {pacote} já instalado")
        except ImportError:
            print(f"   ⬇️  Instalando {pacote}...")
            subprocess.check_call([__import__('sys').executable, "-m", "pip", "install", pacote, "-q"])
            print(f"   ✅ {pacote} instalado")

def extrair_texto_pdf(caminho_pdf):
    """Extrai texto de um PDF"""
    try:
        import pdfplumber
        
        texto_completo = ""
        num_paginas = 0
        
        with pdfplumber.open(caminho_pdf) as pdf:
            num_paginas = len(pdf.pages)
            for pagina in pdf.pages[:5]:  # Apenas primeiras 5 páginas
                texto = pagina.extract_text()
                if texto:
                    texto_completo += texto + "\n"
        
        return texto_completo, num_paginas
    
    except Exception as e:
        return "", 0

def extrair_metadados(arquivo_pdf, texto):
    """Extrai metadados do PDF"""
    
    metadados = {
        'data_sessao': None,
        'tipo_sessao': None,
        'numero_sessao': None,
        'tipo_ata': None,
        'eh_circunstanciada': False
    }
    
    if not texto:
        return metadados
    
    texto_upper = texto.upper()
    
    # Detectar tipo de ata
    if 'CIRCUNSTANCIADA' in texto_upper:
        metadados['tipo_ata'] = 'Circunstanciada'
        metadados['eh_circunstanciada'] = True
    elif 'SUCINTA' in texto_upper:
        metadados['tipo_ata'] = 'Sucinta'
    
    # Detectar tipo de sessão
    if 'EXTRAORDINÁRIA' in texto_upper:
        metadados['tipo_sessao'] = 'Extraordinária'
    elif 'SOLENE' in texto_upper:
        metadados['tipo_sessao'] = 'Solene'
    elif 'ORDINÁRIA' in texto_upper:
        metadados['tipo_sessao'] = 'Ordinária'
    
    # Extrair número de sessão
    padroes_sessao = [
        r'(\d+)(?:ª|a)\s+(?:SESSÃO|SESSÃO)',
        r'SESSÃO\s+(?:Nº|N°|N)\s*(\d+)',
        r'(?:Nº|N°|N)\s*(\d+)\s+(?:SESSÃO|SESSÃO)'
    ]
    
    for padrao in padroes_sessao:
        matches = re.findall(padrao, texto_upper)
        if matches:
            metadados['numero_sessao'] = matches[0]
            break
    
    # Extrair data (padrão: DD de MMMM de YYYY)
    padroes_data = [
        r'(\d{1,2})\s+de\s+([a-záéíóúâêôãõç]+)\s+de\s+(\d{4})',
        r'(\d{1,2})/(\d{1,2})/(\d{4})',
        r'(\d{4})-(\d{1,2})-(\d{1,2})'
    ]
    
    for padrao in padroes_data:
        matches = re.findall(padrao, texto, re.IGNORECASE)
        if matches:
            metadados['data_sessao'] = str(matches[0])
            break
    
    return metadados

def main():
    """Função principal"""
    print("\n" + "="*70)
    print("FASE 2: ENRIQUECIMENTO DE METADADOS - 346 DCLs")
    print("="*70)
    
    # Instalar dependências
    instalar_dependencias()
    
    # Listar todos os PDFs
    print("\n📁 LISTANDO ARQUIVOS PDF")
    todos_pdfs = sorted(DIR_DOWNLOADS.glob('*.pdf'))
    
    print(f"✅ {len(todos_pdfs)} PDFs encontrados")
    
    if not todos_pdfs:
        print("❌ Nenhum PDF encontrado")
        return
    
    # Processar sequencialmente
    print(f"\n📊 PROCESSANDO {len(todos_pdfs)} DCLs COM ENRIQUECIMENTO")
    print("="*70 + "\n")
    
    resultados = []
    
    for i, arquivo in enumerate(todos_pdfs, 1):
        # Extrair texto
        texto, num_paginas = extrair_texto_pdf(arquivo)
        
        # Extrair metadados
        metadados = extrair_metadados(arquivo, texto)
        
        # Extrair informações do nome do arquivo
        nome = arquivo.stem
        partes = nome.split('-')
        
        ano = int(partes[1]) if len(partes) > 1 else 0
        mes = int(partes[2]) if len(partes) > 2 else 0
        tamanho_kb = arquivo.stat().st_size / 1024
        
        resultado = {
            'numero': i,
            'arquivo': arquivo.name,
            'ano': ano,
            'mes': mes,
            'tamanho_kb': tamanho_kb,
            'num_paginas': num_paginas,
            'data_sessao': metadados.get('data_sessao'),
            'tipo_sessao': metadados.get('tipo_sessao'),
            'numero_sessao': metadados.get('numero_sessao'),
            'tipo_ata': metadados.get('tipo_ata'),
            'eh_circunstanciada': metadados.get('eh_circunstanciada'),
            'status': 'processado'
        }
        
        resultados.append(resultado)
        
        if i % 50 == 0:
            print(f"[{i:3d}/{len(todos_pdfs)}] ✅ Processados")
    
    print(f"[{len(todos_pdfs):3d}/{len(todos_pdfs)}] ✅ Processados")
    
    # Salvar JSON
    print(f"\n💾 SALVANDO RESULTADOS")
    print("="*70)
    
    with open(ARQUIVO_SAIDA_JSON, 'w', encoding='utf-8') as f:
        json.dump(resultados, f, ensure_ascii=False, indent=2)
    
    print(f"✅ JSON salvo em: {ARQUIVO_SAIDA_JSON}")
    
    # Salvar Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Enriquecido"
        
        # Cabeçalhos
        headers = ['Nº', 'Arquivo', 'Ano', 'Mês', 'Páginas', 'Data Sessão', 
                   'Tipo Sessão', 'Nº Sessão', 'Tipo Ata', 'Circunstanciada', 'Status']
        ws.append(headers)
        
        # Estilo do cabeçalho
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Dados
        for resultado in resultados:
            eh_circ = "Sim" if resultado.get('eh_circunstanciada') else "Não"
            ws.append([
                resultado.get('numero', ''),
                resultado.get('arquivo', ''),
                resultado.get('ano', ''),
                resultado.get('mes', ''),
                resultado.get('num_paginas', ''),
                resultado.get('data_sessao', ''),
                resultado.get('tipo_sessao', ''),
                resultado.get('numero_sessao', ''),
                resultado.get('tipo_ata', ''),
                eh_circ,
                resultado.get('status', '')
            ])
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 12
        
        # Salvar
        wb.save(ARQUIVO_SAIDA_XLSX)
        
        print(f"✅ Excel salvo em: {ARQUIVO_SAIDA_XLSX}")
    
    except Exception as e:
        print(f"⚠️  Erro ao salvar Excel: {e}")
    
    # Estatísticas
    print(f"\n📊 ESTATÍSTICAS")
    print("="*70)
    
    total = len(resultados)
    circunstanciadas = sum(1 for r in resultados if r.get('eh_circunstanciada'))
    com_data = sum(1 for r in resultados if r.get('data_sessao'))
    com_tipo_sessao = sum(1 for r in resultados if r.get('tipo_sessao'))
    
    print(f"   Total:                  {total}")
    print(f"   Circunstanciadas:       {circunstanciadas}")
    print(f"   Com data extraída:      {com_data}")
    print(f"   Com tipo de sessão:     {com_tipo_sessao}")
    
    print(f"\n✅ FASE 2 CONCLUÍDA COM SUCESSO!")

if __name__ == "__main__":
    main()

