"""
FASE 2: ENRIQUECIMENTO DE METADADOS
Extrai informações detalhadas de cada DCL:
- Data real da sessão
- Tipo de sessão
- Número da sessão
- Tipo de ata
- Paginação
"""

import os
import json
import re
from pathlib import Path
from datetime import datetime
import pdfplumber
import openpyxl
from extrair_data_melhorado import extrair_data_melhorado

DOWNLOADS_DIR = "downloads_2007"
RELATORIO_ENTRADA = "relatorio_dcls_2007.xlsx"
RELATORIO_SAIDA = "relatorio_dcls_2007_enriquecido.xlsx"
DADOS_JSON = "dados_dcls_2007_enriquecidos.json"

def extrair_metadados_dcl(pdf_path):
    """Extrai metadados de um DCL"""
    
    metadados = {
        "data_sessao": None,
        "tipo_sessao": None,
        "numero_sessao": None,
        "tipo_ata": None,
        "paginas_totais": 0,
        "atas_encontradas": []
    }
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            metadados["paginas_totais"] = len(pdf.pages)
            
            # Extrair data real
            data_sessao = extrair_data_melhorado(pdf_path, verbose=False)
            if data_sessao:
                metadados["data_sessao"] = data_sessao
            
            # Procurar por padrões de tipo de sessão
            for page_num in range(min(3, len(pdf.pages))):
                text = pdf.pages[page_num].extract_text()
                if not text:
                    continue
                
                text_upper = text.upper()
                
                # Detectar tipo de sessão
                if "EXTRAORDINÁRIA" in text_upper:
                    metadados["tipo_sessao"] = "Extraordinária"
                elif "SOLENE" in text_upper:
                    metadados["tipo_sessao"] = "Solene"
                elif "ORDINÁRIA" in text_upper:
                    metadados["tipo_sessao"] = "Ordinária"
                
                # Detectar tipo de ata
                if "CIRCUNSTANCIADA" in text_upper:
                    metadados["tipo_ata"] = "Circunstanciada"
                elif "SUCINTA" in text_upper:
                    metadados["tipo_ata"] = "Sucinta"
                
                # Procurar por número de sessão
                pattern = r'(\d+)(?:ª|a)\s+(?:SESSÃO|SESSÃO)'
                matches = re.findall(pattern, text_upper)
                if matches:
                    metadados["numero_sessao"] = matches[0]
    
    except Exception as e:
        print(f"⚠️  Erro ao processar {os.path.basename(pdf_path)}: {e}")
    
    return metadados

def enriquecer_relatorio():
    """Enriquece relatório com metadados"""
    
    print(f"📊 Enriquecendo relatório com metadados...\n")
    
    # Carregar relatório original
    wb = openpyxl.load_workbook(RELATORIO_ENTRADA)
    ws = wb.active
    
    # Adicionar colunas
    ws.cell(row=1, column=8).value = "Data Sessão"
    ws.cell(row=1, column=9).value = "Tipo Sessão"
    ws.cell(row=1, column=10).value = "Nº Sessão"
    ws.cell(row=1, column=11).value = "Tipo Ata"
    ws.cell(row=1, column=12).value = "Páginas"
    
    # Dados para JSON
    dados_json = []
    
    # Processar cada DCL
    dcl_files = sorted([f for f in os.listdir(DOWNLOADS_DIR) if f.endswith('.pdf')])
    
    for idx, filename in enumerate(dcl_files, 2):
        filepath = os.path.join(DOWNLOADS_DIR, filename)
        
        print(f"[{idx-1}/{len(dcl_files)}] {filename}...", end=" ")
        
        # Extrair metadados
        metadados = extrair_metadados_dcl(filepath)
        
        # Adicionar ao Excel
        ws.cell(row=idx, column=8).value = metadados.get("data_sessao", "")
        ws.cell(row=idx, column=9).value = metadados.get("tipo_sessao", "")
        ws.cell(row=idx, column=10).value = metadados.get("numero_sessao", "")
        ws.cell(row=idx, column=11).value = metadados.get("tipo_ata", "")
        ws.cell(row=idx, column=12).value = metadados.get("paginas_totais", "")
        
        # Adicionar ao JSON
        dados_json.append({
            "filename": filename,
            "metadados": metadados
        })
        
        print("✅")
    
    # Salvar Excel
    wb.save(RELATORIO_SAIDA)
    print(f"\n✅ Relatório enriquecido: {RELATORIO_SAIDA}")
    
    # Salvar JSON
    with open(DADOS_JSON, 'w', encoding='utf-8') as f:
        json.dump(dados_json, f, indent=2, ensure_ascii=False)
    print(f"✅ Dados JSON: {DADOS_JSON}")

if __name__ == "__main__":
    enriquecer_relatorio()

