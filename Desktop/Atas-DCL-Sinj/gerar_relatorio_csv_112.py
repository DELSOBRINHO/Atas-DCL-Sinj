#!/usr/bin/env python3
import json
import csv
from pathlib import Path

USUARIO = "omega"
ARQUIVO_JSON = Path(f"C:/Users/{USUARIO}/Desktop/Atas-DCL-Sinj/fase2_atas_2007_final.json")
ARQUIVO_CSV = Path(f"C:/Users/{USUARIO}/Desktop/Atas-DCL-Sinj/relatorio_validacao_112_atas_FINAL.csv")
ARQUIVO_XLSX = Path(f"C:/Users/{USUARIO}/Desktop/Atas-DCL-Sinj/relatorio_validacao_112_atas_FINAL.xlsx")

print("\n" + "="*70)
print("GERAR RELATÓRIO COM 112 ATAS")
print("="*70)

with open(ARQUIVO_JSON, 'r', encoding='utf-8') as f:
    atas = json.load(f)

print(f"\nTotal de atas: {len(atas)}")

# Gerar CSV com encoding UTF-8-SIG (BOM) para compatibilidade com Excel
with open(ARQUIVO_CSV, 'w', newline='', encoding='utf-8-sig') as f:
    writer = csv.writer(f, delimiter=';')

    # Cabeçalho
    writer.writerow([
        'Sessao',
        'Tipo',
        'Data',
        'Pag_Inicio',
        'Pag_Fim',
        'DCL_Original',
        'Nomenclatura'
    ])

    # Dados
    for ata in atas:
        writer.writerow([
            ata['sessao_num'],
            ata['tipo_sessao'],
            ata['data_real'],
            ata['pag_inicio'],
            ata['pag_fim'],
            ata['dcl_original'],
            ata['nomenclatura']
        ])

print(f"✅ Relatório CSV gerado: {ARQUIVO_CSV.name}")

# Tentar gerar XLSX também
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Atas"

    # Cabeçalho
    headers = ['Sessao', 'Tipo', 'Data', 'Pag_Inicio', 'Pag_Fim', 'DCL_Original', 'Nomenclatura']
    ws.append(headers)

    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Dados
    for ata in atas:
        ws.append([
            ata['sessao_num'],
            ata['tipo_sessao'],
            ata['data_real'],
            ata['pag_inicio'],
            ata['pag_fim'],
            ata['dcl_original'],
            ata['nomenclatura']
        ])

    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 35

    wb.save(ARQUIVO_XLSX)
    print(f"✅ Relatório XLSX gerado: {ARQUIVO_XLSX.name}")
except Exception as e:
    print(f"⚠️  Erro ao gerar XLSX: {e}")

print(f"\n" + "="*70)

