#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

script_dir = os.path.dirname(os.path.abspath(__file__))
json_path = os.path.join(script_dir, 'fase2_atas_2007_final.json')

print("📖 Carregando arquivo JSON...")
with open(json_path, 'r', encoding='utf-8') as f:
    atas = json.load(f)

print(f"✅ {len(atas)} atas carregadas")

# Criar workbook para v1.7
print("📊 Criando relatório v1.7...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Conferência Manual"

# Cabeçalhos
headers = [
    'Sessão', 'Tipo', 'Data Real', 'Data Publicação Ata', 'Pág Início', 'Pág Fim', 
    'DCL Original', 'Nomenclatura', 'Validado', 'Observação', 'Ações'
]
ws.append(headers)

# Formatar cabeçalho
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Adicionar dados ordenados
atas_sorted = sorted(atas, key=lambda x: (x['tipo_sessao'], int(str(x['sessao_num']))))
for ata in atas_sorted:
    dcl_original = ata.get('dcl_original', '')
    data_pub = ata.get('data_publicacao_ata', 'N/A')
    
    ws.append([
        str(ata['sessao_num']),
        ata['tipo_sessao'],
        ata['data_real'],
        data_pub,
        ata['pag_inicio'],
        ata['pag_fim'],
        dcl_original,
        ata['nomenclatura'],
        '',  # Validado
        '',  # Observação
        ''   # Ações
    ])

# Ajustar largura das colunas
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 28
ws.column_dimensions['H'].width = 35
ws.column_dimensions['I'].width = 12
ws.column_dimensions['J'].width = 30
ws.column_dimensions['K'].width = 30

# Adicionar bordas
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=11):
    for cell in row:
        cell.border = thin_border
        if cell.row > 1:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Congelar primeira linha
ws.freeze_panes = "A2"

# Salvar arquivo versionado v1.7
version = "v1.7"
data_hoje = datetime.now().strftime("%Y-%m-%d")
output_filename = f"{version}_{data_hoje}.xlsx"
output_path = os.path.join(script_dir, 'documentacao', 'relatorios_conferencia', output_filename)

print(f"💾 Salvando relatório...")
wb.save(output_path)
print(f'✅ Relatório v1.7 gerado: {output_filename}')
print(f'Total de atas: {len(atas)}')
print(f'Correções v1.7:')
print(f'  - Data de publicação extraída do cabeçalho real dos DCLs')
print(f'  - Exemplo: DCL_2007-03-044.pdf → 08/03/2007 (não 01/03/2007)')

