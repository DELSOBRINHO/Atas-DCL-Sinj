#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PROCESSAR DCLs COM OCR - VERSÃO RÁPIDA (APENAS NOVOS)
======================================================

Processa apenas os 327 DCLs novos (com padrão DCL_YYYY-MM-NNNNNNNNNN.pdf)

Uso:
    python processar_ocr_rapido_2007_2008.py

Autor: Sistema de Automação CLDF
Data: 2025-12-22
"""

import os
import sys
import json
import time
from pathlib import Path
import re

# ======================================================================
# CONFIGURAÇÕES
# ======================================================================

USUARIO = "omega"
DIR_DOWNLOADS = Path(f"C:/Users/{USUARIO}/Desktop/Atas-DCL-Sinj/downloads_2007")
DIR_LINKS = Path(f"C:/Users/{USUARIO}/Desktop/Atas-DCL-Sinj/links_2007")
ARQUIVO_LINKS = DIR_LINKS / "dcls_2007.json"
ARQUIVO_SAIDA = Path(f"C:/Users/{USUARIO}/Desktop/Atas-DCL-Sinj/relatorio_dcls_2007_2008.json")

# ======================================================================
# FUNÇÕES PRINCIPAIS
# ======================================================================

def instalar_dependencias():
    """Instala dependências necessárias"""
    print("📦 Verificando dependências...")
    
    import subprocess
    
    pacotes = {
        "pdfplumber": "pdfplumber",
        "openpyxl": "openpyxl"
    }
    
    for modulo, pacote in pacotes.items():
        try:
            __import__(modulo)
            print(f"   ✅ {pacote} já instalado")
        except ImportError:
            print(f"   ⬇️  Instalando {pacote}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", pacote, "-q"])
            print(f"   ✅ {pacote} instalado")

def extrair_texto_pdf(caminho_pdf):
    """Extrai texto de um PDF usando pdfplumber"""
    try:
        import pdfplumber
        
        texto_completo = ""
        num_paginas = 0
        
        with pdfplumber.open(caminho_pdf) as pdf:
            num_paginas = len(pdf.pages)
            for pagina in pdf.pages:
                texto = pagina.extract_text()
                if texto:
                    texto_completo += texto + "\n"
        
        return texto_completo, num_paginas
    
    except Exception as e:
        return "", 0

def detectar_ata_circunstanciada(texto):
    """Detecta se o PDF contém ata circunstanciada"""
    if not texto:
        return False
    
    padroes = [
        r'ata\s+circunstanciada',
        r'circunstanciada',
        r'AC\s*-\s*Ata\s+Circunstanciada',
        r'Ata\s+Circunstanciada'
    ]
    
    texto_lower = texto.lower()
    
    for padrao in padroes:
        if re.search(padrao, texto_lower, re.IGNORECASE):
            return True
    
    return False

def processar_dcl(arquivo_pdf):
    """Processa um DCL individual"""
    try:
        # Extrair texto
        texto, num_paginas = extrair_texto_pdf(arquivo_pdf)
        
        # Detectar ata circunstanciada
        eh_circunstanciada = detectar_ata_circunstanciada(texto)
        
        # Extrair informações do nome do arquivo
        nome = arquivo_pdf.stem
        partes = nome.split('-')
        
        ano = int(partes[1]) if len(partes) > 1 else 0
        mes = int(partes[2]) if len(partes) > 2 else 0
        
        return {
            'arquivo': arquivo_pdf.name,
            'ano': ano,
            'mes': mes,
            'num_paginas': num_paginas,
            'eh_circunstanciada': eh_circunstanciada,
            'tamanho_kb': arquivo_pdf.stat().st_size / 1024,
            'status': 'processado'
        }
    
    except Exception as e:
        return {
            'arquivo': arquivo_pdf.name,
            'status': 'erro',
            'erro': str(e)
        }

def carregar_links():
    """Carrega links do arquivo JSON"""
    if not ARQUIVO_LINKS.exists():
        return {}
    
    try:
        with open(ARQUIVO_LINKS, 'r', encoding='utf-8') as f:
            links = json.load(f)
        
        links_dict = {}
        for link in links:
            try:
                partes = link['url'].split('DCL%20n%C2%BA%20')
                if len(partes) > 1:
                    numero = partes[1].split('%20')[0]
                    ano = link['ano']
                    mes = link['mes']
                    chave = f"DCL_{ano}-{mes:02d}-{numero:03d}"
                    links_dict[chave] = link['url']
            except:
                pass
        
        return links_dict
    
    except Exception as e:
        return {}

def salvar_resultados(resultados, links_dict):
    """Salva resultados em JSON e Excel"""
    print(f"\n💾 SALVANDO RESULTADOS")
    print("="*70)
    
    # Adicionar URLs aos resultados
    for resultado in resultados:
        arquivo = resultado['arquivo']
        chave = arquivo.replace('.pdf', '')
        resultado['url'] = links_dict.get(chave, '')
    
    # Salvar JSON
    with open(ARQUIVO_SAIDA, 'w', encoding='utf-8') as f:
        json.dump(resultados, f, ensure_ascii=False, indent=2)
    
    print(f"✅ Resultados salvos em: {ARQUIVO_SAIDA}")
    
    # Salvar Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "DCLs"
        
        # Cabeçalhos
        headers = ['Nº', 'Arquivo', 'Ano', 'Mês', 'Páginas', 'Circunstanciada', 'Tamanho (KB)', 'Status', 'URL']
        ws.append(headers)
        
        # Estilo do cabeçalho
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Dados
        for i, resultado in enumerate(resultados, 1):
            eh_circ = "Sim" if resultado.get('eh_circunstanciada') else "Não"
            ws.append([
                i,
                resultado.get('arquivo', ''),
                resultado.get('ano', ''),
                resultado.get('mes', ''),
                resultado.get('num_paginas', ''),
                eh_circ,
                f"{resultado.get('tamanho_kb', 0):.0f}",
                resultado.get('status', ''),
                resultado.get('url', '')
            ])
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 50
        
        # Salvar
        arquivo_excel = ARQUIVO_SAIDA.parent / "relatorio_dcls_2007_2008.xlsx"
        wb.save(arquivo_excel)
        
        print(f"✅ Relatório Excel salvo em: {arquivo_excel}")
    
    except Exception as e:
        print(f"⚠️  Erro ao salvar Excel: {e}")

def main():
    """Função principal"""
    print("\n" + "="*70)
    print("PROCESSAR DCLs COM OCR - 2007 + JAN-FEV 2008 (VERSÃO RÁPIDA)")
    print("="*70)
    
    # Instalar dependências
    instalar_dependencias()
    
    # Carregar links
    print("\n📂 CARREGANDO LINKS")
    links_dict = carregar_links()
    print(f"✅ {len(links_dict)} links carregados")
    
    # Listar apenas os 327 novos PDFs (com padrão DCL_YYYY-MM-NNNNNNNNNN.pdf)
    print("\n📁 LISTANDO ARQUIVOS PDF NOVOS")
    todos_pdfs = sorted(DIR_DOWNLOADS.glob('*.pdf'))
    
    # Filtrar apenas os novos (com 10 dígitos no final)
    arquivos_pdf = [f for f in todos_pdfs if re.match(r'DCL_\d{4}-\d{2}-\d{10}\.pdf', f.name)]
    
    print(f"✅ {len(arquivos_pdf)} PDFs novos encontrados (de {len(todos_pdfs)} totais)")
    
    if not arquivos_pdf:
        print("❌ Nenhum PDF novo encontrado em downloads_2007/")
        return
    
    # Processar sequencialmente
    print(f"\n📊 PROCESSANDO {len(arquivos_pdf)} DCLs COM OCR")
    print("="*70 + "\n")
    
    resultados = []
    
    for i, arquivo in enumerate(arquivos_pdf, 1):
        resultado = processar_dcl(arquivo)
        resultados.append(resultado)
        
        # Mostrar progresso
        status = "✅" if resultado.get('status') == 'processado' else "❌"
        eh_circ = "🔵" if resultado.get('eh_circunstanciada') else "⚪"
        
        print(f"[{i:3d}/{len(arquivos_pdf)}] {status} {eh_circ} {arquivo.name}")
    
    # Salvar resultados
    salvar_resultados(resultados, links_dict)
    
    # Estatísticas
    print(f"\n📊 ESTATÍSTICAS")
    print("="*70)
    
    total = len(resultados)
    processados = sum(1 for r in resultados if r.get('status') == 'processado')
    circunstanciadas = sum(1 for r in resultados if r.get('eh_circunstanciada'))
    erros = sum(1 for r in resultados if r.get('status') == 'erro')
    
    print(f"   Total:              {total}")
    print(f"   Processados:        {processados}")
    print(f"   Circunstanciadas:   {circunstanciadas}")
    print(f"   Erros:              {erros}")
    
    print(f"\n✅ PROCESSAMENTO CONCLUÍDO!")

if __name__ == "__main__":
    main()

